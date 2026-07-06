# ============================================================
#  REMI Solar Services — AI Email Triage
#  Created by GOVERNYS (Niels Tilch — CTO)
# ============================================================
"""
app.py — REMI Solar Services email-triage MVP (Flask + SQLite + bcrypt).

Run:
    python init_db.py        # first time only
    python app.py            # then open http://127.0.0.1:5000
"""
import os
import io
import uuid
import sqlite3
import functools
import datetime as dt
import bcrypt
import pyotp
from werkzeug.utils import secure_filename
from flask import (
    Flask, g, render_template, request, redirect, url_for,
    session, flash, abort, send_file,
)
import mistral_client
from ids import uuid7

DB_PATH = os.path.join(os.path.dirname(__file__), "remi_solar.db")
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "storage", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Which department a category is routed to. Tune freely.
CATEGORY_TO_DEPT = {
    "lead": "Sales", "installation": "Sales", "maintenance": "IT",
    "invoices": "Finance", "support": "HR",
}

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-only-change-me")

# --------------------------------------------------------------------------
#  Navigation model — which pages exist, their section, icon, and who sees them.
#  The left menu and the route guards both read from this single source.
# --------------------------------------------------------------------------
ALL = {"global_administrator", "department_viewer", "audit_administrator",
       "classification_verification_administrator", "email_administrator"}

NAV = [
    ("Overview", [
        ("home", "Home", "layout-dashboard", ALL),
    ]),
    ("Mail", [
        ("mailbox", "Department Mailbox", "inbox", {"department_viewer", "global_administrator"}),
        ("reception", "Reception Storage", "mail-warning", {"email_administrator", "global_administrator"}),
    ]),
    ("Operations", [
        ("verification", "Verification", "check-circle-2",
         {"classification_verification_administrator", "global_administrator"}),
        ("audit", "Audit", "scroll-text", {"audit_administrator", "global_administrator"}),
    ]),
    ("Tools", [
        ("test", "System Test", "flask-conical", {"global_administrator"}),
    ]),
    ("Administration", [
        ("users", "Users & Roles", "users", {"global_administrator"}),
        ("departments", "Departments", "building-2", {"global_administrator"}),
        ("settings", "Settings", "settings", {"global_administrator"}),
        ("retention", "Data Retention", "trash-2", {"global_administrator"}),
    ]),
]

# page key -> allowed roles (flattened, for the route guard)
PAGE_ROLES = {key: roles for _, items in NAV for key, _, _, roles in items}
PAGE_ROLES["profile"] = ALL


# --------------------------------------------------------------------------
#  Database helpers
# --------------------------------------------------------------------------
def db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON;")
    return g.db


@app.teardown_appcontext
def close_db(_):
    conn = g.pop("db", None)
    if conn is not None:
        conn.close()


def log_action(action_type, entity_type=None, entity_id=None, email_id=None, details=None):
    db().execute(
        """INSERT INTO audit_logs(uuid, email_id, user_id, action_type, entity_type, entity_id, details, ip_address)
           VALUES (?,?,?,?,?,?,?,?)""",
        (uuid7(), email_id, session.get("user_id"), action_type, entity_type, entity_id, details, request.remote_addr),
    )
    db().commit()


# --------------------------------------------------------------------------
#  Auth + RBAC
# --------------------------------------------------------------------------
def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return db().execute(
        """SELECT u.*, r.name AS role_name, d.name AS dept_name
           FROM users u JOIN roles r ON u.role_id = r.id
           LEFT JOIN departments d ON u.department_id = d.id
           WHERE u.id = ?""", (uid,)
    ).fetchone()


def login_required(view):
    @functools.wraps(view)
    def wrapped(*a, **kw):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return view(*a, **kw)
    return wrapped


def page_guard(page_key):
    """Decorator: only roles allowed for `page_key` may enter."""
    def deco(view):
        @functools.wraps(view)
        @login_required
        def wrapped(*a, **kw):
            user = current_user()
            if user["role_name"] not in PAGE_ROLES.get(page_key, set()):
                abort(403)
            return view(*a, **kw)
        return wrapped
    return deco


@app.context_processor
def inject_nav():
    """Build the role-filtered menu for the base template."""
    user = current_user()
    if not user:
        return {"user": None, "nav": [], "active": request.endpoint}
    role = user["role_name"]
    visible = []
    for section, items in NAV:
        allowed = [(k, label, icon) for k, label, icon, roles in items if role in roles]
        if allowed:
            visible.append((section, allowed))
    pending = db().execute("SELECT COUNT(*) c FROM review_tasks WHERE status='pending'").fetchone()["c"]
    now = dt.datetime.now().strftime("%a · %d %b %Y · %H:%M")
    return {"user": user, "nav": visible, "active": request.endpoint,
            "pending_reviews": pending, "now": now}


# --------------------------------------------------------------------------
#  Routes — auth
# --------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("home"))
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        pw = request.form.get("password", "")
        row = db().execute("SELECT * FROM users WHERE email=? AND is_active=1", (email,)).fetchone()
        if row and bcrypt.checkpw(pw.encode(), row["password_hash"].encode()):
            if row["mfa_enabled"]:
                # Password OK — defer login until the 6-digit code is verified.
                session.clear()
                session["pending_uid"] = row["id"]
                return redirect(url_for("login_mfa"))
            _complete_login(row["id"])
            return redirect(url_for("home"))
        flash("Wrong email or password.", "error")
    return render_template("login.html")


@app.route("/login/mfa", methods=["GET", "POST"])
def login_mfa():
    uid = session.get("pending_uid")
    if not uid:
        return redirect(url_for("login"))
    if request.method == "POST":
        code = request.form.get("code", "").strip().replace(" ", "")
        row = db().execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        if row and pyotp.TOTP(row["mfa_secret"]).verify(code, valid_window=1):
            session.pop("pending_uid", None)
            _complete_login(uid)
            return redirect(url_for("home"))
        flash("Invalid or expired code. Try again.", "error")
    return render_template("login_mfa.html")


def _complete_login(uid):
    session["user_id"] = uid
    db().execute("UPDATE users SET last_login_at=? WHERE id=?",
                 (dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), uid))
    db().commit()
    log_action("login", "users", uid)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# --------------------------------------------------------------------------
#  Routes — pages
# --------------------------------------------------------------------------
@app.route("/")
@login_required
def home():
    d = db()
    user = current_user()
    today = dt.date.today().strftime("%Y-%m-%d")
    is_viewer = user["role_name"] == "department_viewer"
    dep = user["department_id"]
    settings = {r["key"]: r["value"] for r in d.execute("SELECT * FROM app_settings")}

    if is_viewer:
        # Everything is restricted to emails routed to this viewer's department.
        scope = "JOIN routing_assignments r ON r.email_id=e.id AND r.department_id=?"
        sp = (dep,)
        scope_label = (user["dept_name"] or "your") + " department"
        stats = {
            "today": d.execute(f"SELECT COUNT(*) c FROM emails e {scope} WHERE date(e.received_at)=?",
                               (dep, today)).fetchone()["c"],
            "total": d.execute(f"SELECT COUNT(*) c FROM emails e {scope}", sp).fetchone()["c"],
            "auto":  d.execute(f"""SELECT COUNT(*) c FROM emails e {scope}
                                   JOIN email_analysis a ON a.email_id=e.id AND a.is_manual=0""", sp).fetchone()["c"],
            "high":  d.execute(f"""SELECT COUNT(*) c FROM emails e {scope}
                                   JOIN email_analysis a ON a.email_id=e.id
                                   JOIN priorities p ON a.priority_id=p.id
                                   WHERE p.level IN ('high','critical')""", sp).fetchone()["c"],
        }
        recent = d.execute(f"""
            SELECT e.id, e.sender_email, e.subject, e.status,
                   c.name AS category, p.level AS priority, a.confidence, ? AS dept
            FROM emails e {scope}
            LEFT JOIN email_analysis a ON a.email_id=e.id
            LEFT JOIN email_categories c ON a.category_id=c.id
            LEFT JOIN priorities p ON a.priority_id=p.id
            ORDER BY e.received_at DESC LIMIT 8
        """, (user["dept_name"], dep)).fetchall()
        by_cat = d.execute(f"""
            SELECT c.name, COUNT(*) n FROM emails e {scope}
            JOIN email_analysis a ON a.email_id=e.id
            JOIN email_categories c ON a.category_id=c.id
            GROUP BY c.name ORDER BY n DESC
        """, sp).fetchall()
        # viewers can't access the verification queue, so show their high-priority items
        attention_title = "Needs attention"
        attention = [{"danger": r["priority"] == "critical", "line1": (r["priority"] or "").title(),
                      "line2": r["subject"]}
                     for r in d.execute(f"""
            SELECT e.subject, p.level AS priority FROM emails e {scope}
            JOIN email_analysis a ON a.email_id=e.id
            JOIN priorities p ON a.priority_id=p.id
            WHERE p.level IN ('high','critical') ORDER BY a.priority_score DESC LIMIT 5
        """, sp).fetchall()]
    else:
        scope_label = "company-wide"
        stats = {
            "today": d.execute("SELECT COUNT(*) c FROM emails WHERE date(received_at)=?", (today,)).fetchone()["c"],
            "total": d.execute("SELECT COUNT(*) c FROM emails").fetchone()["c"],
            "pending": d.execute("SELECT COUNT(*) c FROM review_tasks WHERE status='pending'").fetchone()["c"],
            "auto": d.execute("SELECT COUNT(*) c FROM email_analysis WHERE is_manual=0").fetchone()["c"],
        }
        recent = d.execute("""
            SELECT e.id, e.sender_email, e.subject, e.status,
                   c.name AS category, p.level AS priority, a.confidence, dep.name AS dept
            FROM emails e
            LEFT JOIN email_analysis a ON a.email_id = e.id
            LEFT JOIN email_categories c ON a.category_id = c.id
            LEFT JOIN priorities p ON a.priority_id = p.id
            LEFT JOIN routing_assignments r ON r.email_id = e.id
            LEFT JOIN departments dep ON r.department_id = dep.id
            ORDER BY e.received_at DESC LIMIT 8
        """).fetchall()
        by_cat = d.execute("""
            SELECT c.name, COUNT(*) n FROM email_analysis a
            JOIN email_categories c ON a.category_id=c.id GROUP BY c.name ORDER BY n DESC
        """).fetchall()
        attention_title = "Needs review"
        attention = [{"danger": r["reason"] == "low_confidence",
                      "line1": r["reason"].replace("_", " ").title() + (
                          " · %.2f" % r["confidence"] if r["confidence"] else ""),
                      "line2": r["subject"]}
                     for r in d.execute("""
            SELECT t.reason, e.subject, a.confidence FROM review_tasks t
            JOIN emails e ON t.email_id=e.id
            LEFT JOIN email_analysis a ON t.analysis_id=a.id
            WHERE t.status='pending' ORDER BY t.created_at DESC LIMIT 5
        """).fetchall()]

    return render_template("home.html", stats=stats, recent=recent, by_cat=by_cat,
                           attention=attention, attention_title=attention_title,
                           is_viewer=is_viewer, scope_label=scope_label, settings=settings)


@app.route("/mailbox")
@page_guard("mailbox")
def mailbox():
    user = current_user()
    # Global admin can see all; viewer sees only their department.
    if user["role_name"] == "global_administrator":
        dept_filter, params = "", ()
        title = "All Departments"
    else:
        dept_filter, params = "WHERE r.department_id = ?", (user["department_id"],)
        title = user["dept_name"]
    rows = db().execute(f"""
        SELECT e.id, e.uuid, e.sender_email, e.subject, e.received_at, e.status,
               c.name AS category, p.level AS priority, a.confidence, a.summary
        FROM emails e
        JOIN routing_assignments r ON r.email_id = e.id
        LEFT JOIN email_analysis a ON a.email_id = e.id
        LEFT JOIN email_categories c ON a.category_id = c.id
        LEFT JOIN priorities p ON a.priority_id = p.id
        {dept_filter}
        ORDER BY e.received_at DESC
    """, params).fetchall()
    return render_template("mailbox.html", emails=rows, title=title)


@app.route("/email/<euid>")
@login_required
def email_detail(euid):
    user = current_user()
    if user["role_name"] not in {"global_administrator", "department_viewer",
                                 "classification_verification_administrator"}:
        abort(403)
    e = db().execute("SELECT * FROM emails WHERE uuid=?", (euid,)).fetchone()
    if not e:
        abort(404)
    eid = e["id"]
    analysis = db().execute("""
        SELECT a.*, c.name AS category, p.level AS priority, rt.name AS request_type
        FROM email_analysis a
        LEFT JOIN email_categories c ON a.category_id=c.id
        LEFT JOIN priorities p ON a.priority_id=p.id
        LEFT JOIN request_types rt ON a.request_type_id=rt.id
        WHERE a.email_id=? ORDER BY a.created_at DESC LIMIT 1
    """, (eid,)).fetchone()
    fields = db().execute("SELECT * FROM extracted_fields WHERE email_id=?", (eid,)).fetchall()
    atts = db().execute("SELECT * FROM attachments WHERE email_id=?", (eid,)).fetchall()
    log_action("access", "emails", eid, email_id=eid)
    return render_template("email_detail.html", e=e, analysis=analysis, fields=fields, atts=atts)


@app.route("/verification")
@page_guard("verification")
def verification():
    rows = db().execute("""
        SELECT t.id, t.reason, t.status, t.created_at, e.id AS email_id, e.uuid AS email_uuid,
               e.sender_email, e.subject, a.confidence, c.name AS category, p.level AS priority
        FROM review_tasks t
        JOIN emails e ON t.email_id=e.id
        LEFT JOIN email_analysis a ON t.analysis_id=a.id
        LEFT JOIN email_categories c ON a.category_id=c.id
        LEFT JOIN priorities p ON a.priority_id=p.id
        WHERE t.status='pending' ORDER BY a.confidence ASC
    """).fetchall()
    return render_template("verification.html", tasks=rows)


@app.route("/verification/<int:task_id>/<decision>", methods=["POST"])
@page_guard("verification")
def verification_decision(task_id, decision):
    if decision not in {"confirmed", "corrected", "rejected"}:
        abort(400)
    db().execute(
        "UPDATE review_tasks SET status=?, reviewed_by=?, reviewed_at=? WHERE id=?",
        (decision, session["user_id"], dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), task_id),
    )
    db().commit()
    log_action("manual_override", "review_tasks", task_id, details=f'{{"decision":"{decision}"}}')
    flash(f"Task marked {decision}.", "success")
    return redirect(url_for("verification"))


@app.route("/audit")
@page_guard("audit")
def audit():
    action = request.args.get("action", "")
    per_page = 20
    page = max(1, request.args.get("page", 1, type=int))

    where = "WHERE l.action_type=?" if action else ""
    params = (action,) if action else ()
    total = db().execute(f"SELECT COUNT(*) c FROM audit_logs l {where}", params).fetchone()["c"]
    pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, pages)
    offset = (page - 1) * per_page

    rows = db().execute(f"""
        SELECT l.*, u.full_name AS who, e.subject
        FROM audit_logs l
        LEFT JOIN users u ON l.user_id=u.id
        LEFT JOIN emails e ON l.email_id=e.id
        {where}
        ORDER BY l.created_at DESC
        LIMIT ? OFFSET ?
    """, (*params, per_page, offset)).fetchall()

    actions = [r["action_type"] for r in db().execute("SELECT DISTINCT action_type FROM audit_logs")]
    start = 0 if total == 0 else offset + 1
    end = offset + len(rows)
    return render_template("audit.html", logs=rows, actions=actions, current=action,
                           page=page, pages=pages, total=total, start=start, end=end)


@app.route("/audit/export")
@page_guard("audit")
def audit_export():
    """Download the audit log (optionally filtered by action) as an .xlsx file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    action = request.args.get("action", "")
    base = """
        SELECT l.uuid, l.created_at, l.action_type, u.full_name AS who, u.email AS who_email,
               l.entity_type, l.entity_id, l.model_version, e.subject, l.details, l.ip_address
        FROM audit_logs l
        LEFT JOIN users u ON l.user_id=u.id
        LEFT JOIN emails e ON l.email_id=e.id
    """
    if action:
        rows = db().execute(base + " WHERE l.action_type=? ORDER BY l.created_at DESC", (action,)).fetchall()
    else:
        rows = db().execute(base + " ORDER BY l.created_at DESC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit log"
    headers = ["Log ID (uuid)", "Timestamp", "Action", "Actor", "Actor email", "Entity",
               "Entity ID", "AI model", "Email subject", "Details", "IP address"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="0E1424")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
    for r in rows:
        ws.append([r["uuid"], r["created_at"], r["action_type"], r["who"] or "system", r["who_email"] or "",
                   r["entity_type"] or "", r["entity_id"] if r["entity_id"] is not None else "",
                   r["model_version"] or "", r["subject"] or "", r["details"] or "", r["ip_address"] or ""])
    widths = [38, 20, 16, 20, 26, 16, 10, 22, 34, 40, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_action("access", "audit_logs", details='{"action":"export_xlsx"}')
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M")
    name = f"remi-solar-audit{('-' + action) if action else ''}-{stamp}.xlsx"
    return send_file(buf, as_attachment=True, download_name=name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/reception")
@page_guard("reception")
def reception():
    rows = db().execute("""
        SELECT e.*, (SELECT COUNT(*) FROM attachments WHERE email_id=e.id) AS n_att
        FROM emails e ORDER BY e.received_at DESC
    """).fetchall()
    return render_template("reception.html", emails=rows)


@app.route("/users")
@page_guard("users")
def users():
    per_page = 10
    page = max(1, request.args.get("page", 1, type=int))
    total = db().execute("SELECT COUNT(*) c FROM users").fetchone()["c"]
    pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, pages)
    offset = (page - 1) * per_page
    rows = db().execute("""
        SELECT u.*, r.name AS role_name, d.name AS dept_name
        FROM users u JOIN roles r ON u.role_id=r.id
        LEFT JOIN departments d ON u.department_id=d.id
        ORDER BY u.full_name LIMIT ? OFFSET ?
    """, (per_page, offset)).fetchall()
    roles = db().execute("SELECT * FROM roles").fetchall()
    depts = db().execute("SELECT * FROM departments").fetchall()
    start = 0 if total == 0 else offset + 1
    end = offset + len(rows)
    return render_template("users.html", users=rows, roles=roles, depts=depts,
                           page=page, pages=pages, total=total, start=start, end=end)


@app.route("/users/create", methods=["POST"])
@page_guard("users")
def users_create():
    f = request.form
    pw_hash = bcrypt.hashpw(f["password"].encode(), bcrypt.gensalt()).decode()
    # Front-end sends the department UUID; resolve it to the internal id.
    dept_uuid = f.get("department_uuid") or ""
    dept_row = db().execute("SELECT id FROM departments WHERE uuid=?", (dept_uuid,)).fetchone() if dept_uuid else None
    try:
        db().execute(
            "INSERT INTO users(uuid, full_name, email, password_hash, role_id, department_id) VALUES (?,?,?,?,?,?)",
            (uuid7(), f["full_name"], f["email"].lower(), pw_hash, f["role_id"],
             dept_row["id"] if dept_row else None),
        )
        db().commit()
        log_action("access", "users", details='{"action":"create_user"}')
        flash("User created.", "success")
    except sqlite3.IntegrityError:
        flash("That email is already in use.", "error")
    return redirect(url_for("users"))


@app.route("/departments")
@page_guard("departments")
def departments():
    rows = db().execute("""
        SELECT d.*, (SELECT COUNT(*) FROM users WHERE department_id=d.id) AS n_users,
               (SELECT COUNT(*) FROM routing_assignments WHERE department_id=d.id) AS n_mail
        FROM departments d ORDER BY d.name
    """).fetchall()
    return render_template("departments.html", depts=rows)


@app.route("/departments/create", methods=["POST"])
@page_guard("departments")
def departments_create():
    try:
        db().execute("INSERT INTO departments(uuid, name, email_address, description) VALUES (?,?,?,?)",
                     (uuid7(), request.form["name"], request.form.get("email_address") or None,
                      request.form.get("description")))
        db().commit()
        flash("Department added.", "success")
    except sqlite3.IntegrityError:
        flash("That department name or email already exists.", "error")
    return redirect(url_for("departments"))


@app.route("/settings", methods=["GET", "POST"])
@page_guard("settings")
def settings():
    if request.method == "POST":
        for key, val in request.form.items():
            db().execute("UPDATE app_settings SET value=? WHERE key=?", (val, key))
        db().commit()
        log_action("access", "app_settings", details='{"action":"update_settings"}')
        flash("Settings saved.", "success")
        return redirect(url_for("settings"))

    cfg = {r["key"]: r["value"] for r in db().execute("SELECT * FROM app_settings ORDER BY key")}
    u = db().execute("""
        SELECT
          COALESCE(SUM(pages),0)         AS pages,
          COALESCE(SUM(input_tokens),0)  AS in_tok,
          COALESCE(SUM(output_tokens),0) AS out_tok,
          COALESCE(SUM(CASE WHEN kind='classification' THEN 1 END),0) AS emails,
          COALESCE(SUM(CASE WHEN kind='ocr' THEN 1 END),0)            AS ocr_docs,
          AVG(CASE WHEN kind='classification' THEN confidence END)    AS avg_conf
        FROM ai_usage_events
    """).fetchone()
    usage = {"pages": u["pages"], "in_tok": u["in_tok"], "out_tok": u["out_tok"],
             "emails": u["emails"], "ocr_docs": u["ocr_docs"],
             "avg_conf": u["avg_conf"] or 0, "total_tok": (u["in_tok"] or 0) + (u["out_tok"] or 0)}
    by_model = db().execute("""
        SELECT model_version, provider, kind, COUNT(*) AS calls,
               COALESCE(SUM(input_tokens+output_tokens),0) AS tokens,
               COALESCE(SUM(pages),0) AS pages
        FROM ai_usage_events GROUP BY model_version, kind ORDER BY tokens DESC
    """).fetchall()
    return render_template("settings.html", cfg=cfg, usage=usage, by_model=by_model)


@app.route("/retention")
@page_guard("retention")
def retention():
    policies = db().execute("""
        SELECT rp.*, c.name AS category FROM retention_policies rp
        LEFT JOIN email_categories c ON rp.category_id=c.id ORDER BY rp.id
    """).fetchall()
    deletions = db().execute("""
        SELECT dr.*, u.full_name AS who FROM deletion_records dr
        LEFT JOIN users u ON dr.performed_by=u.id ORDER BY dr.id DESC LIMIT 50
    """).fetchall()
    return render_template("retention.html", policies=policies, deletions=deletions)


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = current_user()
    if request.method == "POST":
        cur = request.form["current"]
        new = request.form["new"]
        if not bcrypt.checkpw(cur.encode(), user["password_hash"].encode()):
            flash("Current password is incorrect.", "error")
        elif len(new) < 8:
            flash("New password must be at least 8 characters.", "error")
        else:
            h = bcrypt.hashpw(new.encode(), bcrypt.gensalt()).decode()
            db().execute("UPDATE users SET password_hash=? WHERE id=?", (h, user["id"]))
            db().commit()
            flash("Password updated.", "success")
        return redirect(url_for("profile"))

    # MFA setup state: a secret staged in the session means setup is in progress.
    setup_secret = session.get("mfa_setup_secret")
    otp_uri = None
    if setup_secret and not user["mfa_enabled"]:
        otp_uri = pyotp.totp.TOTP(setup_secret).provisioning_uri(
            name=user["email"], issuer_name="REMI Solar Services")
    return render_template("profile.html", setup_secret=setup_secret, otp_uri=otp_uri)


@app.route("/profile/mfa/init", methods=["POST"])
@login_required
def mfa_init():
    session["mfa_setup_secret"] = pyotp.random_base32()
    return redirect(url_for("profile"))


@app.route("/profile/mfa/cancel", methods=["POST"])
@login_required
def mfa_cancel():
    session.pop("mfa_setup_secret", None)
    return redirect(url_for("profile"))


@app.route("/profile/mfa/confirm", methods=["POST"])
@login_required
def mfa_confirm():
    user = current_user()
    secret = session.get("mfa_setup_secret")
    code = request.form.get("code", "").strip().replace(" ", "")
    if secret and pyotp.TOTP(secret).verify(code, valid_window=1):
        db().execute("UPDATE users SET mfa_enabled=1, mfa_secret=? WHERE id=?", (secret, user["id"]))
        db().commit()
        session.pop("mfa_setup_secret", None)
        log_action("access", "users", user["id"], details='{"action":"mfa_enabled"}')
        flash("Two-factor authentication is now on.", "success")
    else:
        flash("That code didn't match. Scan the QR again and retry.", "error")
    return redirect(url_for("profile"))


@app.route("/profile/mfa/disable", methods=["POST"])
@login_required
def mfa_disable():
    user = current_user()
    code = request.form.get("code", "").strip().replace(" ", "")
    if user["mfa_secret"] and pyotp.TOTP(user["mfa_secret"]).verify(code, valid_window=1):
        db().execute("UPDATE users SET mfa_enabled=0, mfa_secret=NULL WHERE id=?", (user["id"],))
        db().commit()
        log_action("access", "users", user["id"], details='{"action":"mfa_disabled"}')
        flash("Two-factor authentication turned off.", "success")
    else:
        flash("Enter a valid current code to turn off two-factor.", "error")
    return redirect(url_for("profile"))


# --------------------------------------------------------------------------
#  System Test — global admin sends a synthetic email through the pipeline
# --------------------------------------------------------------------------
@app.route("/test", methods=["GET", "POST"])
@page_guard("test")
def test():
    if request.method == "GET":
        return render_template("test.html", result=None)

    d = db()
    sender = request.form.get("sender", "").strip() or "test@example.com"
    subject = request.form.get("subject", "").strip()
    body = request.form.get("body", "").strip()
    cfg = {r["key"]: r["value"] for r in d.execute("SELECT * FROM app_settings")}
    threshold = float(cfg.get("confidence_threshold", "0.70") or 0.70)
    flagged = [t.strip().lower() for t in (cfg.get("flagged_topics", "") or "").split(",") if t.strip()]
    ocr_provider = cfg.get("ocr_provider", "local")
    llm_provider = cfg.get("llm_provider", "local")

    files = [f for f in request.files.getlist("attachments") if f and f.filename]

    # 1. create the email
    mid = f"<test-{uuid.uuid4().hex[:10]}@remi>"
    euid = uuid7()
    cur = d.execute("""INSERT INTO emails(uuid, message_id, sender_email, sender_name, subject, body,
                                          received_at, status, has_attachments, storage_ref)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (euid, mid, sender, sender.split("@")[0], subject, body,
                     dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     "processing", 1 if files else 0, "secure://test"))
    eid = cur.lastrowid
    log_action("ingestion", "emails", eid, email_id=eid, details='{"source":"system_test"}')

    # 2. attachments + OCR
    ocr_text, ocr_rows = "", []
    edir = os.path.join(UPLOAD_DIR, str(eid))
    os.makedirs(edir, exist_ok=True)
    for f in files:
        name = secure_filename(f.filename)
        data = f.read()
        with open(os.path.join(edir, name), "wb") as out:
            out.write(data)
        ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
        ftype = "pdf" if ext == "pdf" else "image" if ext in ("png", "jpg", "jpeg", "gif", "webp") \
            else "docx" if ext in ("doc", "docx") else "other"
        acur = d.execute("""INSERT INTO attachments(email_id, filename, file_type, file_size, storage_ref)
                            VALUES (?,?,?,?,?)""", (eid, name, ftype, len(data), f"secure://test/{eid}/{name}"))
        aid = acur.lastrowid
        ocr = mistral_client.ocr_document(data, f.mimetype or "", name, provider=ocr_provider)
        d.execute("""INSERT INTO ocr_results(attachment_id, model_version, extracted_text, confidence, status)
                     VALUES (?,?,?,?,?)""",
                  (aid, ocr["model_version"], ocr["text"], ocr["confidence"], "success"))
        _record_usage(d, eid, "ocr", ocr)
        ocr_text += "\n" + ocr["text"]
        ocr_rows.append({"filename": name, "type": ftype, "size": len(data),
                         "text": ocr["text"], "model": ocr["model_version"], "provider": ocr["provider"]})
        log_action("ocr_extraction", "attachments", aid, email_id=eid, details=ocr["model_version"])

    # 3. classify
    result = mistral_client.classify_email(subject, body, ocr_text, provider=llm_provider)
    cat_id = lookup_id(d, "email_categories", "name", result["category"])
    prio_id = lookup_id(d, "priorities", "level", result["priority"])
    rtype_id = lookup_id(d, "request_types", "name", result["request_type"])
    acur = d.execute("""INSERT INTO email_analysis(email_id, category_id, request_type_id, priority_id,
                                                   priority_score, confidence, summary, model_version)
                        VALUES (?,?,?,?,?,?,?,?)""",
                     (eid, cat_id, rtype_id, prio_id, result["priority_score"],
                      result["confidence"], result["summary"], result["model_version"]))
    analysis_id = acur.lastrowid
    _record_usage(d, eid, "classification", result)
    log_action("classification", "email_analysis", analysis_id, email_id=eid,
               details=f'{{"category":"{result["category"]}","confidence":{result["confidence"]}}}')

    for k, v in (result.get("fields") or {}).items():
        d.execute("""INSERT INTO extracted_fields(email_id, analysis_id, field_name, field_value, model_version)
                     VALUES (?,?,?,?,?)""", (eid, analysis_id, k, str(v), result["model_version"]))

    # 4. route OR send to manual review
    blob = f"{subject} {body} {ocr_text}".lower()
    hit_topic = next((t for t in flagged if t and t in blob), None)
    routed_to, review_reason = None, None
    if result["confidence"] < threshold:
        review_reason = "low_confidence"
    elif hit_topic:
        review_reason = "flagged_topic"

    if review_reason:
        d.execute("INSERT INTO review_tasks(email_id, analysis_id, reason) VALUES (?,?,?)",
                  (eid, analysis_id, review_reason))
        d.execute("UPDATE emails SET status='manual_review' WHERE id=?", (eid,))
    else:
        dept_name = CATEGORY_TO_DEPT.get(result["category"])
        dept = d.execute("SELECT * FROM departments WHERE name=?", (dept_name,)).fetchone() if dept_name else None
        d.execute("""INSERT INTO routing_assignments(email_id, analysis_id, department_id, priority_id)
                     VALUES (?,?,?,?)""", (eid, analysis_id, dept["id"] if dept else None, prio_id))
        d.execute("UPDATE emails SET status='routed' WHERE id=?", (eid,))
        routed_to = {"dept": dept["name"], "email": dept["email_address"]} if dept else \
            {"dept": "Default mailbox", "email": "unclassified@remi-solar.eu"}
        log_action("routing", "emails", eid, email_id=eid,
                   details=f'{{"dept":"{routed_to["dept"]}"}}')

    d.commit()
    view = {
        "email_id": eid, "email_uuid": euid, "message_id": mid, "sender": sender, "subject": subject,
        "received_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "result": result, "ocr": ocr_rows, "attachments_detected": bool(files),
        "review_reason": review_reason, "routed_to": routed_to,
        "threshold": threshold, "flagged_topic": hit_topic,
        "ocr_provider": ocr_provider, "llm_provider": llm_provider,
    }
    return render_template("test.html", result=view)


def _record_usage(d, email_id, kind, payload):
    d.execute(
        """INSERT INTO ai_usage_events(email_id, kind, provider, model_version,
                                       pages, input_tokens, output_tokens, confidence)
           VALUES (?,?,?,?,?,?,?,?)""",
        (email_id, kind, payload.get("provider"), payload.get("model_version"),
         payload.get("pages", 0), payload.get("input_tokens", 0),
         payload.get("output_tokens", 0), payload.get("confidence")),
    )


def lookup_id(d, table, col, value):
    if not value:
        return None
    row = d.execute(f"SELECT id FROM {table} WHERE {col}=?", (value,)).fetchone()
    return row["id"] if row else None


@app.errorhandler(403)
def forbidden(_):
    return render_template("error.html", code=403,
                           msg="You don't have access to this page."), 403


@app.errorhandler(404)
def not_found(_):
    return render_template("error.html", code=404, msg="Page not found."), 404


if __name__ == "__main__":
    if not os.path.exists(DB_PATH):
        print("No database found. Run:  python init_db.py")
    app.run(debug=True, port=5000)
